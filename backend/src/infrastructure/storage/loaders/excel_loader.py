import csv
from pathlib import Path
from typing import List
import openpyxl
from backend.src.infrastructure.storage.loaders.base_loader import BaseDocumentLoader, ExtractedChunk


class ExcelCsvDocumentLoader(BaseDocumentLoader):
    """Loader strategy for extracting tabular text from CSV and XLSX files."""

    def load(self, file_path: Path) -> List[ExtractedChunk]:
        ext = file_path.suffix.lower()
        if ext == ".csv":
            return self._load_csv(file_path)
        elif ext == ".xlsx":
            return self._load_xlsx(file_path)
        return []

    def _load_csv(self, file_path: Path) -> List[ExtractedChunk]:
        lines = []
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            for row in reader:
                if any(cell.strip() for cell in row):
                    lines.append(", ".join(row))

        content = "\n".join(lines)
        if not content.strip():
            return []

        return [ExtractedChunk(content=content, page_number=1, metadata={"row_count": len(lines)})]

    def _load_xlsx(self, file_path: Path) -> List[ExtractedChunk]:
        wb = openpyxl.load_workbook(str(file_path), data_only=True)
        chunks: List[ExtractedChunk] = []

        for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
            sheet = wb[sheet_name]
            sheet_rows = []
            for row in sheet.iter_rows(values_only=True):
                str_cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if str_cells:
                    sheet_rows.append(" | ".join(str_cells))

            sheet_content = "\n".join(sheet_rows)
            if sheet_content.strip():
                chunks.append(
                    ExtractedChunk(
                        content=f"Sheet: {sheet_name}\n" + sheet_content,
                        page_number=sheet_idx,
                        metadata={"sheet_name": sheet_name},
                    )
                )

        return chunks
